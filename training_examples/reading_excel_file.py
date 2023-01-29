import openpyxl

file = "C:/Users/rusgl/Desktop/random.xlsx"
workbook = openpyxl.load_workbook(file)
sheet = workbook.active
sheet.cell(2, 3).value = "Haha"
rows = sheet.max_row  # count max number of rows in a excel sheet 6
columns = sheet.max_column  # count number of columns in a excel sheet
data = sheet.cell(2, 3).value

for row in range(1, rows+1):
    for col in range(1, columns+1):
        print(sheet.cell(row, col).value, end="       ")
    print()
print(data)
workbook.save(file)
